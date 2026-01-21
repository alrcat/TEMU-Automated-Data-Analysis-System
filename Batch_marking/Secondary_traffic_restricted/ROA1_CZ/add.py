# -*- coding: utf-8 -*-
"""
批量标记二级流量受限商品
"""

import os
import sys
import pymysql
import pandas as pd
import openpyxl
from datetime import datetime

sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))))
from config import DEFAULT_DB_CONFIG


def get_table_name_from_dir():
    """从当前目录名获取表名"""
    current_dir = os.path.basename(os.path.dirname(__file__))
    return current_dir


def get_excel_goods_ids(directory):
    """从目录中所有Excel文件获取Goods ID列表（第3列）"""
    goods_ids = set()
    excel_files = [f for f in os.listdir(directory) if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$')]
    
    for excel_file in excel_files:
        excel_path = os.path.join(directory, excel_file)
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            ws = wb.active
            
            header_row_idx = None
            for i, row in enumerate(ws.iter_rows(max_row=20)):
                values = [str(cell.value) if cell.value else '' for cell in row[:10]]
                if 'Goods ID' in values:
                    header_row_idx = i
                    break
            
            if header_row_idx is not None:
                for row in ws.iter_rows(min_row=header_row_idx + 3):
                    goods_id_cell = row[2]
                    if goods_id_cell.value:
                        goods_id = str(goods_id_cell.value).strip()
                        if goods_id and goods_id != 'Goods ID':
                            goods_ids.add(goods_id)
            else:
                for row in ws.iter_rows(min_row=3):
                    goods_id_cell = row[2]
                    if goods_id_cell.value:
                        goods_id = str(goods_id_cell.value).strip()
                        if goods_id:
                            goods_ids.add(goods_id)
        except Exception as e:
            print(f"读取文件 {excel_file} 时出错: {e}")
    
    return goods_ids


def check_column_exists(cursor, table_name, column_name):
    """检查表中是否存在指定列"""
    query = f"""
    SELECT COUNT(*) 
    FROM INFORMATION_SCHEMA.COLUMNS 
    WHERE TABLE_SCHEMA = DATABASE() 
      AND TABLE_NAME = %s 
      AND COLUMN_NAME = %s
    """
    cursor.execute(query, (table_name, column_name))
    return cursor.fetchone()[0] > 0

def get_status_not_null_goods_ids(table_name, db_config):
    """从数据库表中获取Status不为空的goods_id，去重（如果Status列不存在则返回所有goods_id）"""
    conn = pymysql.connect(**db_config)
    try:
        cursor = conn.cursor()
        
        if check_column_exists(cursor, table_name, 'Status'):
            query = f"""
            SELECT DISTINCT goods_id
            FROM `{table_name}`
            WHERE Status IS NOT NULL
            """
        else:
            query = f"""
            SELECT DISTINCT goods_id
            FROM `{table_name}`
            """
        
        cursor.execute(query)
        results = cursor.fetchall()
        goods_ids = {str(row[0]) for row in results}
        return goods_ids
    finally:
        cursor.close()
        conn.close()


def update_reason_for_goods_ids(table_name, goods_ids, target_date, db_config):
    """更新指定goods_id和date_label的Reason字段为'Secondary_traffic_restricted'，如果目标日期不存在则使用最近的date_label"""
    if not goods_ids:
        return 0
    
    conn = pymysql.connect(**db_config)
    updated_count = 0
    try:
        cursor = conn.cursor()
        
        for goods_id in goods_ids:
            date_to_update = target_date
            
            check_query = f"""
            SELECT COUNT(*) 
            FROM `{table_name}` 
            WHERE goods_id = %s AND date_label = %s
            """
            cursor.execute(check_query, (goods_id, target_date))
            exists = cursor.fetchone()[0] > 0
            
            if not exists:
                nearest_query = f"""
                SELECT date_label 
                FROM `{table_name}` 
                WHERE goods_id = %s AND date_label <= %s
                ORDER BY date_label DESC 
                LIMIT 1
                """
                cursor.execute(nearest_query, (goods_id, target_date))
                result = cursor.fetchone()
                if result:
                    date_to_update = result[0]
                    if isinstance(date_to_update, datetime):
                        date_to_update = date_to_update.strftime('%Y-%m-%d')
                    elif hasattr(date_to_update, 'strftime'):
                        date_to_update = date_to_update.strftime('%Y-%m-%d')
                    else:
                        date_to_update = str(date_to_update)
                else:
                    continue
            
            update_query = f"""
            UPDATE `{table_name}`
            SET Reason = 'Secondary_traffic_restricted'
            WHERE goods_id = %s AND date_label = %s
            """
            cursor.execute(update_query, (goods_id, date_to_update))
            updated_count += cursor.rowcount
        
        conn.commit()
        return updated_count
    finally:
        cursor.close()
        conn.close()


def save_history(current_dir, goods_ids, target_date, updated_count):
    """保存历史记录到txt文件"""
    history_dir = os.path.join(current_dir, '历史记录')
    os.makedirs(history_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'Secondary_traffic_restricted_{target_date}_{timestamp}.txt'
    filepath = os.path.join(history_dir, filename)
    
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(f"标记日期: {target_date}\n")
        f.write(f"执行时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"成功标记 {len(goods_ids)} 个goods_id，更新了 {updated_count} 条记录\n\n")
        f.write(f"已标记的goods_id列表: {', '.join(sorted(goods_ids))}\n")
    
    print(f"历史记录已保存到: {filepath}")


def main():
    """主函数"""
    current_dir = os.path.dirname(__file__)
    table_name = get_table_name_from_dir()
    target_date = '2025-12-16'
    
    print(f"表名: {table_name}")
    print(f"目标日期: {target_date}")
    
    db_config = DEFAULT_DB_CONFIG.copy()
    
    print("正在查询数据库中Status不为空的goods_id...")
    db_goods_ids = get_status_not_null_goods_ids(table_name, db_config)
    print(f"找到 {len(db_goods_ids)} 个不重复的goods_id")
    
    print("正在读取Excel文件中的Goods ID...")
    excel_goods_ids = get_excel_goods_ids(current_dir)
    print(f"Excel文件中找到 {len(excel_goods_ids)} 个不重复的Goods ID")
    
    matching_goods_ids = db_goods_ids & excel_goods_ids
    matching_goods_ids_sorted = sorted(matching_goods_ids)
    print(f"匹配的goods_id数量: {len(matching_goods_ids)}")
    
    if matching_goods_ids:
        print(f"\n正在更新数据库...")
        updated_count = update_reason_for_goods_ids(table_name, matching_goods_ids, target_date, db_config)
        print(f"成功标记 {len(matching_goods_ids)} 个goods_id，更新了 {updated_count} 条记录")
        print(f"已标记的goods_id列表: {', '.join(matching_goods_ids_sorted)}")
        save_history(current_dir, matching_goods_ids, target_date, updated_count)
    else:
        print("\n没有需要更新的记录")


if __name__ == '__main__':
    main()

