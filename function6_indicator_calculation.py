# -*- coding: utf-8 -*-
"""
功能6：指标计算模块
根据数据库数据和本地xlsx文件计算各项指标并生成图表
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import os
import json
from datetime import datetime, timedelta
from flask import jsonify, request
from db_utils import get_db_connection
from config import (
    load_config, save_config, get_current_table, get_db_config,
    DEFAULT_DB_CONFIG, DEFAULT_SALES_DB_CONFIG,
    DEFAULT_PALLET_DB_CONFIG, DEFAULT_PRODUCT_DB_CONFIG
)
from plot_utils import plot_to_base64


def get_eastern_europe_time():
    """
    获取东欧时间（当前时间减去7小时）
    返回: datetime对象
    """
    return datetime.now() - timedelta(hours=7)


def normalize_goods_id(goods_id_val):
    """
    标准化goods_id格式
    处理浮点数（如12345.0）、科学计数法、字符串等各种格式
    返回: 标准化的字符串格式goods_id（只保留数字字符），如果无效则返回None
    """
    if goods_id_val is None or pd.isna(goods_id_val):
        return None
    
    goods_id_str = str(goods_id_val).strip()
    
    # 过滤无效值
    if not goods_id_str or goods_id_str.lower() in ['nan', 'none', '']:
        return None
    
    # 去除所有非数字字符，只保留数字（这是最激进但最可靠的方法）
    digits_only = ''.join(c for c in goods_id_str if c.isdigit())
    
    if not digits_only:
        return None
    
    # 对于纯数字字符串，直接返回（已经是最标准格式）
    return digits_only


def load_indicator_config():
    """加载指标计算配置文件"""
    config = load_config()
    return config.get('indicator_config', {
        'unpriced_data_dir': '',
        'traffic_restricted_data_dir': ''
    })


# ===== 缓存功能 =====

def get_indicator_cache_file_path(table_name, target_date):
    """
    获取缓存文件路径
    返回: 缓存文件路径
    """
    cache_dir = 'Cache_Indicator'
    if not os.path.exists(cache_dir):
        os.makedirs(cache_dir)
    
    # 文件名格式：表名_日期.json，例如：FR_2025-12-16.json
    # 去掉表名中的ROA1_前缀（如果存在）
    table_suffix = table_name.replace('ROA1_', '')
    filename = f"{table_suffix}_{target_date}.json"
    return os.path.join(cache_dir, filename)


def load_indicator_cache(table_name, target_date):
    """
    加载指标计算缓存
    返回: 缓存数据字典，如果不存在则返回None
    """
    try:
        cache_file = get_indicator_cache_file_path(table_name, target_date)
        if os.path.exists(cache_file):
            with open(cache_file, 'r', encoding='utf-8') as f:
                cache_data = json.load(f)
                return cache_data
        return None
    except Exception as e:
        print(f"加载指标缓存失败: {e}")
        return None


def convert_to_json_serializable(obj):
    """
    递归转换对象为JSON可序列化的类型
    处理 Decimal, set, 等类型
    """
    from decimal import Decimal
    
    if isinstance(obj, Decimal):
        return float(obj)
    elif isinstance(obj, dict):
        return {k: convert_to_json_serializable(v) for k, v in obj.items()}
    elif isinstance(obj, (list, tuple, set)):
        return [convert_to_json_serializable(item) for item in obj]
    elif isinstance(obj, (int, float, str, bool, type(None))):
        return obj
    else:
        # 其他类型转换为字符串
        return str(obj)


def save_indicator_cache(table_name, target_date, results, analysis_time):
    """
    保存指标计算缓存
    参数:
        table_name: 表名
        target_date: 目标日期
        results: 计算结果字典
        analysis_time: 分析耗时
    """
    try:
        cache_file = get_indicator_cache_file_path(table_name, target_date)
        
        # 构建缓存数据（不包含图表，因为base64图片太大）
        cache_data = {
            'table_name': table_name,
            'target_date': target_date,
            'analysis_time': analysis_time,
            'results': {}
        }
        
        # 复制结果，但排除图表数据
        for key, value in results.items():
            if key in ['indicator_8', 'indicator_9']:
                # 图表数据不缓存
                cache_data['results'][key] = {
                    'name': value.get('name', ''),
                    'value': None,  # 图表不缓存
                    'unit': value.get('unit', '')
                }
            else:
                # 转换所有数据为JSON可序列化格式
                cache_data['results'][key] = convert_to_json_serializable(value)
        
        with open(cache_file, 'w', encoding='utf-8') as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=2)
        
        print(f"指标缓存已保存到: {cache_file}")
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"保存指标缓存失败: {e}")


def save_indicator_config(config_data):
    """保存指标计算配置文件"""
    config = load_config()
    config['indicator_config'] = config_data
    return save_config(config)


def read_excel_files_to_goods_ids(unpriced_dir, restricted_dir, site_name):
    """
    从Excel文件读取goods_id数据（实际读取文件）
    返回: (unpriced_goods_ids, restricted_goods_ids, unpriced_file_paths_mtimes, restricted_file_paths_mtimes)
    """
    unpriced_goods_ids = set()
    restricted_goods_ids = set()

    try:
        # 处理未核价数据
        unpriced_site_dir = os.path.join(unpriced_dir, site_name)
        if os.path.exists(unpriced_site_dir):
            unpriced_files = [f for f in os.listdir(unpriced_site_dir) if f.endswith('.xlsx')]

            for file in unpriced_files:
                file_path = os.path.join(unpriced_site_dir, file)
                try:
                    # 尝试读取Template sheet，如果不存在则读取第一个sheet
                    try:
                        df = pd.read_excel(file_path, sheet_name='Template', header=None)
                    except:
                        df = pd.read_excel(file_path, header=None)
                    
                    if len(df) > 2 and len(df.columns) > 2:  # 确保有至少3行3列
                        # 从第3行（索引2）开始，读取第3列（索引2）的所有数据
                        goods_id_column = df.iloc[2:, 2]  # 从第3行开始的所有行的第3列
                        for goods_id_val in goods_id_column:
                            if pd.notna(goods_id_val):
                                goods_id_str = str(goods_id_val).strip()
                                if goods_id_str and goods_id_str.lower() not in ['nan', 'none', '']:
                                    # goods_id可能是单个ID，也可能是多个用空格分隔的ID
                                    # 处理可能的换行符、制表符等
                                    goods_id_str = goods_id_str.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
                                    ids = goods_id_str.split()
                                    for goods_id in ids:
                                        # 使用标准化函数处理goods_id
                                        normalized_id = normalize_goods_id(goods_id)
                                        if normalized_id:
                                            unpriced_goods_ids.add(normalized_id)
                except Exception as e:
                    print(f"读取未核价文件 {file} 出错: {e}")
                    continue

        # 处理限流数据
        restricted_site_dir = os.path.join(restricted_dir, site_name)
        if os.path.exists(restricted_site_dir):
            restricted_files = [f for f in os.listdir(restricted_site_dir) if f.endswith('.xlsx')]

            for file in restricted_files:
                file_path = os.path.join(restricted_site_dir, file)
                try:
                    # 尝试读取Template sheet，如果不存在则读取第一个sheet
                    try:
                        df = pd.read_excel(file_path, sheet_name='Template', header=None)
                    except:
                        df = pd.read_excel(file_path, header=None)
                    
                    if len(df) > 2 and len(df.columns) > 2:  # 确保有至少3行3列
                        # 从第3行（索引2）开始，读取第3列（索引2）的所有数据
                        goods_id_column = df.iloc[2:, 2]  # 从第3行开始的所有行的第3列
                        for goods_id_val in goods_id_column:
                            if pd.notna(goods_id_val):
                                goods_id_str = str(goods_id_val).strip()
                                if goods_id_str and goods_id_str.lower() not in ['nan', 'none', '']:
                                    # goods_id可能是单个ID，也可能是多个用空格分隔的ID
                                    # 处理可能的换行符、制表符等
                                    goods_id_str = goods_id_str.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
                                    ids = goods_id_str.split()
                                    for goods_id in ids:
                                        # 使用标准化函数处理goods_id
                                        normalized_id = normalize_goods_id(goods_id)
                                        if normalized_id:
                                            restricted_goods_ids.add(normalized_id)
                except Exception as e:
                    print(f"读取限流文件 {file} 出错: {e}")
                    continue

        return unpriced_goods_ids, restricted_goods_ids

    except Exception as e:
        print(f"读取Excel文件出错: {e}")
        import traceback
        traceback.print_exc()
        return set(), set()


def validate_excel_directories(unpriced_dir, restricted_dir):
    """
    验证Excel文件目录
    返回: (is_valid, error_message, file_counts)
    """
    try:
        current_table = get_current_table()
        site_name = current_table  # 如 ROA1_FR

        # 检查未核价数据目录
        unpriced_site_dir = os.path.join(unpriced_dir, site_name)
        if not os.path.exists(unpriced_site_dir):
            return False, f"未核价数据目录不存在: {unpriced_site_dir}", None

        unpriced_files = [f for f in os.listdir(unpriced_site_dir) if f.endswith('.xlsx')]
        if len(unpriced_files) == 0:
            return False, f"未核价数据目录中没有xlsx文件: {unpriced_site_dir}", None

        # 检查限流数据目录
        restricted_site_dir = os.path.join(restricted_dir, site_name)
        if not os.path.exists(restricted_site_dir):
            return False, f"限流数据目录不存在: {restricted_site_dir}", None

        restricted_files = [f for f in os.listdir(restricted_site_dir) if f.endswith('.xlsx')]
        if len(restricted_files) == 0:
            return False, f"限流数据目录中没有xlsx文件: {restricted_site_dir}", None

        file_counts = {
            'unpriced_files': len(unpriced_files),
            'restricted_files': len(restricted_files)
        }

        return True, "验证成功", file_counts

    except Exception as e:
        return False, f"验证过程中出错: {str(e)}", None


def get_excel_data(unpriced_dir, restricted_dir):
    """
    获取Excel文件中的goods_id数据
    从第3列（Goods ID列，索引为2）第3行（索引为2）开始读取所有goods_id
    返回: (unpriced_goods_ids, restricted_goods_ids)
    """
    current_table = get_current_table()
    site_name = current_table
    
    try:
        unpriced_goods_ids, restricted_goods_ids = read_excel_files_to_goods_ids(unpriced_dir, restricted_dir, site_name)
        return unpriced_goods_ids, restricted_goods_ids
    except Exception as e:
        print(f"获取Excel数据出错: {e}")
        import traceback
        traceback.print_exc()
        return set(), set()


def get_active_products_data(current_table):
    """
    获取商品表中的在售商品数据
    返回: (active_goods_ids, at_risk_goods_ids)
    """
    _, _, _, product_config = get_db_config()

    try:
        conn = get_db_connection(product_config)
        cursor = conn.cursor()

        # 获取Active状态的商品
        active_query = f"""
        SELECT DISTINCT goods_id
        FROM `{current_table}`
        WHERE detail_status = 'Active'
        """
        cursor.execute(active_query)
        active_results = cursor.fetchall()
        # 使用标准化函数处理goods_id
        active_goods_ids = set()
        for row in active_results:
            normalized_id = normalize_goods_id(row[0])
            if normalized_id:
                active_goods_ids.add(normalized_id)

        # 获取At Risk状态的商品
        at_risk_query = f"""
        SELECT DISTINCT goods_id
        FROM `{current_table}`
        WHERE detail_status = 'At Risk'
        """
        cursor.execute(at_risk_query)
        at_risk_results = cursor.fetchall()
        # 使用标准化函数处理goods_id
        at_risk_goods_ids = set()
        for row in at_risk_results:
            normalized_id = normalize_goods_id(row[0])
            if normalized_id:
                at_risk_goods_ids.add(normalized_id)

        cursor.close()
        conn.close()

        return active_goods_ids, at_risk_goods_ids

    except Exception as e:
        print(f"获取商品数据出错: {e}")
        return set(), set()


def get_sales_data(current_table, sales_table_name, end_date=None):
    """
    获取销售表中的动销商品数据
    参数:
        current_table: 当前表名
        sales_table_name: 销售表名
        end_date: 结束日期（datetime.date对象），如果为None则不过滤日期（获取所有历史动销品）
    返回: sales_goods_ids (set)
    """
    _, sales_config, _, _ = get_db_config()

    try:
        conn = get_db_connection(sales_config)
        cursor = conn.cursor()

        # 获取有动销记录的商品ID
        if end_date is None:
            # 如果没有指定日期，获取所有历史动销品
            sales_query = f"""
            SELECT DISTINCT goods_id
            FROM `{sales_table_name}`
            """
            cursor.execute(sales_query)
        else:
            # 如果指定了日期，只获取该日期及之前的动销品
            if isinstance(end_date, datetime):
                end_date = end_date.date()
            elif isinstance(end_date, str):
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            
            end_date_str = end_date.strftime('%Y-%m-%d')
            sales_query = f"""
            SELECT DISTINCT goods_id
            FROM `{sales_table_name}`
            WHERE date_label <= %s
            """
            cursor.execute(sales_query, (end_date_str,))
        
        sales_results = cursor.fetchall()
        # 使用标准化函数处理goods_id
        sales_goods_ids = set()
        for row in sales_results:
            normalized_id = normalize_goods_id(row[0])
            if normalized_id:
                sales_goods_ids.add(normalized_id)

        cursor.close()
        conn.close()

        return sales_goods_ids

    except Exception as e:
        print(f"获取销售数据出错: {e}")
        return set()


def get_recent_sales_volume(sales_table_name, days=7, end_date=None):
    """
    获取近N天的日均单量
    参数:
        sales_table_name: 销售表名
        days: 天数（默认7天）
        end_date: 结束日期（datetime.date对象），如果为None则使用昨天（东欧时间-1天）
    返回: 平均单量（保留2位小数）
    """
    _, sales_config, _, _ = get_db_config()

    try:
        conn = get_db_connection(sales_config)
        cursor = conn.cursor()

        # 计算N天前的日期
        if end_date is None:
            end_date = (get_eastern_europe_time() - timedelta(days=1)).date()  # 昨天
        elif isinstance(end_date, datetime):
            end_date = end_date.date()
        elif isinstance(end_date, str):
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        start_date = end_date - timedelta(days=days-1)  # N天前

        start_date_str = start_date.strftime('%Y-%m-%d')
        end_date_str = end_date.strftime('%Y-%m-%d')

        # 获取指定日期范围内的总单量
        volume_query = f"""
        SELECT SUM(`Units ordered`) as total_volume
        FROM `{sales_table_name}`
        WHERE date_label >= %s AND date_label <= %s
        """
        cursor.execute(volume_query, (start_date_str, end_date_str))
        result = cursor.fetchone()
        total_volume = result[0] if result[0] else 0

        # 计算日均单量
        avg_daily_volume = round(total_volume / days, 2)

        cursor.close()
        conn.close()

        return avg_daily_volume

    except Exception as e:
        print(f"获取销售量数据出错: {e}")
        return 0.0


def get_gmv_data(sales_table_name, days=30, end_date=None):
    """
    获取近N天的GMV数据
    参数:
        sales_table_name: 销售表名
        days: 天数（默认30天）
        end_date: 结束日期（datetime.date对象），如果为None则使用昨天（东欧时间-1天）
    返回: DataFrame包含日期、销售额、销量
    如果当日没找到数据就默认为0
    """
    _, sales_config, _, _ = get_db_config()

    try:
        conn = get_db_connection(sales_config)
        cursor = conn.cursor()

        # 计算日期范围
        if end_date is None:
            end_date = (get_eastern_europe_time() - timedelta(days=1)).date()  # 昨天
        elif isinstance(end_date, datetime):
            end_date = end_date.date()
        elif isinstance(end_date, str):
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        start_date = end_date - timedelta(days=days-1)  # N天前

        start_date_str = start_date.strftime('%Y-%m-%d')
        end_date_str = end_date.strftime('%Y-%m-%d')

        # 获取每日GMV和销量数据
        gmv_query = f"""
        SELECT
            date_label,
            COALESCE(SUM(`Base price sales`), 0) as daily_gmv,
            COALESCE(SUM(`Units ordered`), 0) as daily_volume
        FROM `{sales_table_name}`
        WHERE date_label >= %s AND date_label <= %s
        GROUP BY date_label
        ORDER BY date_label
        """
        cursor.execute(gmv_query, (start_date_str, end_date_str))
        results = cursor.fetchall()

        # 转换为DataFrame
        if results:
            df = pd.DataFrame(results, columns=['date_label', 'daily_gmv', 'daily_volume'])
            # 将date_label转换为日期（去掉时间部分）
            df['date_label'] = pd.to_datetime(df['date_label']).dt.date
            # 然后再转换为datetime以便后续合并
            df['date_label'] = pd.to_datetime(df['date_label'])
        else:
            # 如果没有数据，创建空DataFrame
            df = pd.DataFrame(columns=['date_label', 'daily_gmv', 'daily_volume'])

        # 生成完整的日期序列（确保所有日期都有数据，只保留日期部分）
        date_range = pd.date_range(start=start_date, end=end_date, freq='D')
        date_df = pd.DataFrame({'date_label': date_range})

        # 合并数据，缺失的日期填充为0
        if len(df) > 0:
            df = date_df.merge(df, on='date_label', how='left')
        else:
            df = date_df.copy()
            df['daily_gmv'] = 0
            df['daily_volume'] = 0

        df['daily_gmv'] = df['daily_gmv'].fillna(0)
        df['daily_volume'] = df['daily_volume'].fillna(0)

        # 确保数据类型正确
        df['daily_gmv'] = pd.to_numeric(df['daily_gmv'], errors='coerce').fillna(0)
        df['daily_volume'] = pd.to_numeric(df['daily_volume'], errors='coerce').fillna(0)

        cursor.close()
        conn.close()

        return df

    except Exception as e:
        print(f"获取GMV数据出错: {e}")
        import traceback
        traceback.print_exc()
        # 即使出错，也返回完整的日期序列（值为0）
        try:
            if end_date is None:
                end_date = (get_eastern_europe_time() - timedelta(days=1)).date()
            elif isinstance(end_date, datetime):
                end_date = end_date.date()
            elif isinstance(end_date, str):
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            start_date = end_date - timedelta(days=days-1)
            date_range = pd.date_range(start=start_date, end=end_date, freq='D')
            df = pd.DataFrame({
                'date_label': date_range,
                'daily_gmv': [0] * len(date_range),
                'daily_volume': [0] * len(date_range)
            })
            return df
        except:
            return pd.DataFrame()


# ===== 指标计算函数 =====

def calculate_non_restricted_active_products(unpriced_goods_ids=None, restricted_goods_ids=None):
    """
    指标1：非限流在售商品数量
    参数: unpriced_goods_ids, restricted_goods_ids - Excel数据（可选，如果为None则从配置读取）
    返回: (数量, goods_id集合)
    """
    # 如果未提供Excel数据，则从配置读取
    if unpriced_goods_ids is None or restricted_goods_ids is None:
        config = load_indicator_config()
        unpriced_dir = config.get('unpriced_data_dir', '')
        restricted_dir = config.get('traffic_restricted_data_dir', '')

        if not unpriced_dir or not restricted_dir:
            return 0, set()

        unpriced_goods_ids, restricted_goods_ids = get_excel_data(unpriced_dir, restricted_dir)

    current_table = get_current_table()

    try:
        # 获取数据库中的在售商品
        active_goods_ids, at_risk_goods_ids = get_active_products_data(current_table)

        # 计算无风险active数据和有风险active数据
        low_risk_active = active_goods_ids
        high_risk_active = at_risk_goods_ids

        # 计算非限流在售商品数量
        # 公式：(无风险active + 有风险active) - 未核价数据 - 限流数据
        all_active = low_risk_active.union(high_risk_active)
        excluded_goods = unpriced_goods_ids.union(restricted_goods_ids)
        non_restricted_active = all_active - excluded_goods

        return len(non_restricted_active), non_restricted_active

    except Exception as e:
        print(f"计算非限流在售商品数量出错: {e}")
        return 0, set()


def calculate_secondary_restriction_ratio(unpriced_goods_ids=None, restricted_goods_ids=None):
    """
    指标2：二次限流占比
    参数: unpriced_goods_ids, restricted_goods_ids - Excel数据（可选，如果为None则从配置读取）
    返回: 百分比值（保留2位小数）
    """
    # 如果未提供Excel数据，则从配置读取
    if restricted_goods_ids is None:
        config = load_indicator_config()
        unpriced_dir = config.get('unpriced_data_dir', '')
        restricted_dir = config.get('traffic_restricted_data_dir', '')

        if not unpriced_dir or not restricted_dir:
            return 0.0

        _, restricted_goods_ids = get_excel_data(unpriced_dir, restricted_dir)

    current_table = get_current_table()

    try:
        restricted_count = len(restricted_goods_ids)

        # 获取数据库中的在售商品
        active_goods_ids, at_risk_goods_ids = get_active_products_data(current_table)
        total_active_count = len(active_goods_ids) + len(at_risk_goods_ids)

        # 计算占比
        if total_active_count > 0:
            ratio = (restricted_count / total_active_count) * 100
            return round(ratio, 2)
        else:
            return 0.0

    except Exception as e:
        print(f"计算二次限流占比出错: {e}")
        return 0.0


def calculate_historical_sales_products():
    """
    指标3：历史动销品数量
    返回: (数量, goods_id集合)
    """
    current_table = get_current_table()
    sales_table_name = f"{current_table}_Sales"

    try:
        sales_goods_ids = get_sales_data(current_table, sales_table_name)
        return len(sales_goods_ids), sales_goods_ids

    except Exception as e:
        print(f"计算历史动销品数量出错: {e}")
        return 0, set()


def calculate_active_sales_products(unpriced_goods_ids=None, restricted_goods_ids=None):
    """
    指标4：在售动销品数量
    参数: unpriced_goods_ids, restricted_goods_ids - Excel数据（可选，如果为None则从配置读取）
    返回: (数量, goods_id集合)
    """
    try:
        # 获取非限流在售商品数据
        _, non_restricted_active_goods = calculate_non_restricted_active_products(unpriced_goods_ids, restricted_goods_ids)

        # 获取历史动销品数据
        _, historical_sales_goods = calculate_historical_sales_products()

        # 计算交集
        active_sales_goods = non_restricted_active_goods.intersection(historical_sales_goods)

        return len(active_sales_goods), active_sales_goods

    except Exception as e:
        print(f"计算在售动销品数量出错: {e}")
        return 0, set()


def calculate_secondary_restriction_sales_ratio(unpriced_goods_ids=None, restricted_goods_ids=None):
    """
    指标5：二次限流动销品占比
    参数: unpriced_goods_ids, restricted_goods_ids - Excel数据（可选，如果为None则从配置读取）
    返回: 百分比值（保留2位小数）
    """
    # 如果未提供Excel数据，则从配置读取
    if restricted_goods_ids is None:
        config = load_indicator_config()
        unpriced_dir = config.get('unpriced_data_dir', '')
        restricted_dir = config.get('traffic_restricted_data_dir', '')

        if not unpriced_dir or not restricted_dir:
            return 0.0

        _, restricted_goods_ids = get_excel_data(unpriced_dir, restricted_dir)

    try:
        # 获取历史动销品数据
        _, historical_sales_goods = calculate_historical_sales_products()

        # 计算历史动销被限流的数据
        restricted_sales_goods = historical_sales_goods.intersection(restricted_goods_ids)
        restricted_sales_count = len(restricted_sales_goods)

        # 计算占比
        historical_sales_count = len(historical_sales_goods)
        if historical_sales_count > 0:
            ratio = (restricted_sales_count / historical_sales_count) * 100
            return round(ratio, 2)
        else:
            return 0.0

    except Exception as e:
        print(f"计算二次限流动销品占比出错: {e}")
        return 0.0


def calculate_average_daily_volume():
    """
    指标6：日均单量（近7日）
    返回: 日均单量（保留2位小数）
    """
    current_table = get_current_table()
    sales_table_name = f"{current_table}_Sales"

    try:
        return get_recent_sales_volume(sales_table_name, days=7)

    except Exception as e:
        print(f"计算日均单量出错: {e}")
        return 0.0


def calculate_process_data(unpriced_goods_ids=None, restricted_goods_ids=None):
    """
    指标7：过程数据展示
    参数: unpriced_goods_ids, restricted_goods_ids - Excel数据（可选，如果为None则从配置读取）
    返回: 包含各项过程数据的字典
    """
    # 如果未提供Excel数据，则从配置读取
    if restricted_goods_ids is None:
        config = load_indicator_config()
        unpriced_dir = config.get('unpriced_data_dir', '')
        restricted_dir = config.get('traffic_restricted_data_dir', '')

        if not unpriced_dir or not restricted_dir:
            return {}

        _, restricted_goods_ids = get_excel_data(unpriced_dir, restricted_dir)

    current_table = get_current_table()

    try:
        # 获取数据库数据
        active_goods_ids, at_risk_goods_ids = get_active_products_data(current_table)

        # 获取历史动销品数据
        _, historical_sales_goods = calculate_historical_sales_products()

        process_data = {
            'low_risk_active_count': len(active_goods_ids),
            'high_risk_active_count': len(at_risk_goods_ids),
            'restricted_data_count': len(restricted_goods_ids),
            'restricted_sales_count': len(historical_sales_goods.intersection(restricted_goods_ids))
        }

        return process_data

    except Exception as e:
        print(f"计算过程数据出错: {e}")
        return {}


def generate_gmv_chart(df, title, bar_color, line_color):
    """
    通用GMV图表生成函数
    参数:
        df: 包含date_label, daily_gmv, daily_volume的DataFrame
        title: 图表标题
        bar_color: 柱状图颜色
        line_color: 折线图颜色
    返回: base64编码的图表字符串
    """
    try:
        if len(df) == 0:
            return None

        # 创建双轴图表（进一步增大尺寸以提高可读性）
        if '近30天' in title or '近7日' in title:
            fig, ax1 = plt.subplots(figsize=(20, 12))
        else:
            fig, ax1 = plt.subplots(figsize=(16, 10))

        # 柱状图：销售额
        ax1.bar(df['date_label'], df['daily_gmv'], alpha=0.7, color=bar_color, label='销售额')
        ax1.set_xlabel('日期', fontsize=16, fontweight='bold', color='black')
        ax1.set_ylabel('销售额', color=bar_color, fontsize=16, fontweight='bold')
        ax1.tick_params(axis='y', labelcolor='black', labelsize=14)
        ax1.tick_params(axis='x', labelcolor='black', labelsize=14)

        # 右轴：销量曲线图
        ax2 = ax1.twinx()
        if title == '近30天GMV图':
            # 30天图使用曲线显示周期趋势
            ax2.plot(df['date_label'], df['daily_volume'], color=line_color, marker='o', linewidth=2.5,
                    markersize=5, linestyle='-', markerfacecolor='white', markeredgewidth=1.5, label='销量')
        else:
            # 7天图保持原有样式
            ax2.plot(df['date_label'], df['daily_volume'], color=line_color, marker='o', linewidth=3,
                    markersize=6, label='销量')

        if '近30天' in title or '近7日' in title:
            ax2.set_ylabel('销量', color=line_color, fontsize=16, fontweight='bold')
            ax2.tick_params(axis='y', labelcolor='black', labelsize=14)
        else:
            ax2.set_ylabel('销量', color=line_color, fontsize=14, fontweight='bold')
            ax2.tick_params(axis='y', labelcolor=line_color, labelsize=12)

        # 设置标题
        if '近30天' in title or '近7日' in title:
            plt.title(title, fontsize=18, fontweight='bold', color='black')
            plt.xticks(rotation=45)
        else:
            plt.title(title, fontsize=16, fontweight='bold')
            plt.xticks(rotation=45)

        # 添加图例（增大字体）
        lines1, labels1 = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        if '近30天' in title or '近7日' in title:
            legend = ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left', fontsize=14, framealpha=0.9)
        else:
            legend = ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left', fontsize=12, framealpha=0.9)
        legend.get_frame().set_facecolor('white')

        # 设置网格线以提高可读性
        ax1.grid(True, alpha=0.3, linestyle='--')

        plt.tight_layout()

        # 转换为base64
        chart_base64 = plot_to_base64(fig)
        return chart_base64

    except Exception as e:
        print(f"生成GMV图表出错: {e}")
        return None


def calculate_30day_gmv_chart():
    """
    指标8：近30天GMV图（兼容旧调用）
    返回: base64编码的图表字符串
    """
    current_table = get_current_table()
    sales_table_name = f"{current_table}_Sales"
    df = get_gmv_data(sales_table_name, days=30)
    return generate_gmv_chart(df, '近30天GMV图', 'skyblue', 'red')


def calculate_7day_gmv_chart():
    """
    指标9：日均GMV图（兼容旧调用）
    返回: base64编码的图表字符串
    """
    current_table = get_current_table()
    sales_table_name = f"{current_table}_Sales"
    df = get_gmv_data(sales_table_name, days=7)
    return generate_gmv_chart(df, '日均GMV图（近7日）', 'lightgreen', 'orange')


# ===== Flask路由函数 =====

def indicator_calculation(target_date=None, use_cache=True):
    """
    主指标计算函数
    参数:
        target_date: 目标日期（字符串格式 'YYYY-MM-DD'），如果为None则使用昨天（东欧时间-1天）
        use_cache: 是否使用缓存，默认True。如果为False，重新计算并覆盖缓存
    返回所有指标的计算结果
    """
    import time
    start_time = time.time()
    
    try:
        # 处理日期参数
        if target_date is None:
            end_date = (get_eastern_europe_time() - timedelta(days=1)).date()
        else:
            if isinstance(target_date, str):
                end_date = datetime.strptime(target_date, '%Y-%m-%d').date()
            elif isinstance(target_date, datetime):
                end_date = target_date.date()
            else:
                end_date = target_date
        
        target_date_str = end_date.strftime('%Y-%m-%d')
        current_table = get_current_table()
        sales_table_name = f"{current_table}_Sales"
        
        # 如果使用缓存，先检查缓存
        if use_cache:
            cache_file_path = get_indicator_cache_file_path(current_table, target_date_str)
            print(f"[缓存检查] 表名: {current_table}, 日期: {target_date_str}")
            print(f"[缓存检查] 缓存文件路径: {cache_file_path}")
            print(f"[缓存检查] 缓存文件是否存在: {os.path.exists(cache_file_path)}")
            
            cache_data = load_indicator_cache(current_table, target_date_str)
            if cache_data:
                # 从缓存加载数据，不生成图表（加快速度）
                print(f"[缓存命中] 直接从缓存返回数据，不生成图表")
                end_time = time.time()
                cache_analysis_time = round(end_time - start_time, 2)
                
                results = cache_data.get('results', {})
                
                # 缓存模式下不生成图表，直接返回None（加快速度）
                results['indicator_8'] = {
                    'name': '近30天GMV图',
                    'value': None,  # 缓存模式不生成图表
                    'unit': ''
                }
                
                results['indicator_9'] = {
                    'name': '日均GMV图',
                    'value': None,  # 缓存模式不生成图表
                    'unit': ''
                }
                
                return jsonify({
                    'success': True,
                    'data': results,
                    'analysis_time': cache_analysis_time,
                    'from_cache': True
                })
            else:
                print(f"[缓存未命中] 缓存文件不存在或加载失败，将重新计算")
        
        results = {}

        # 只读取一次Excel数据
        config = load_indicator_config()
        unpriced_dir = config.get('unpriced_data_dir', '')
        restricted_dir = config.get('traffic_restricted_data_dir', '')
        
        unpriced_goods_ids = set()
        restricted_goods_ids = set()
        
        if unpriced_dir and restricted_dir:
            unpriced_goods_ids, restricted_goods_ids = get_excel_data(unpriced_dir, restricted_dir)

        # ===== 一次性获取所有需要的基础数据 =====
        # 获取商品表数据（只查询一次）
        active_goods_ids, at_risk_goods_ids = get_active_products_data(current_table)
        all_active_goods = active_goods_ids.union(at_risk_goods_ids)
        
        # 获取销售表数据（只查询一次，使用end_date筛选）
        sales_goods_ids = get_sales_data(current_table, sales_table_name, end_date=end_date)
        
        # 计算公共数据
        excluded_goods = unpriced_goods_ids.union(restricted_goods_ids)
        non_restricted_active_goods = all_active_goods - excluded_goods
        restricted_sales_goods = sales_goods_ids.intersection(restricted_goods_ids)

        # ===== 指标计算（使用预加载的数据） =====
        
        # 指标1：非限流在售商品数量
        count_1 = len(non_restricted_active_goods)
        results['indicator_1'] = {
            'name': '非限流在售商品数量',
            'value': count_1,
            'unit': '个'
        }

        # 指标2：二次限流占比
        # 计算公式：限流数据数量 / (非限流在售商品数量 + 限流数据数量)
        restricted_count = len(restricted_goods_ids)
        denominator = count_1 + restricted_count
        if denominator > 0:
            ratio_2 = round((restricted_count / denominator) * 100, 2)
        else:
            ratio_2 = 0.0
        results['indicator_2'] = {
            'name': '二次限流占比',
            'value': ratio_2,
            'unit': '%'
        }

        # 指标3：历史动销品数量
        count_3 = len(sales_goods_ids)
        results['indicator_3'] = {
            'name': '历史动销品数量',
            'value': count_3,
            'unit': '个'
        }

        # 指标4：在售动销品数量
        active_sales_goods = non_restricted_active_goods.intersection(sales_goods_ids)
        count_4 = len(active_sales_goods)
        results['indicator_4'] = {
            'name': '在售动销品数量',
            'value': count_4,
            'unit': '个',
            'goods_ids': sorted(list(active_sales_goods))  # 返回goods_id列表（排序后的列表）
        }

        # 指标5：二次限流动销品占比
        historical_sales_count = len(sales_goods_ids)
        if historical_sales_count > 0:
            ratio_5 = round((len(restricted_sales_goods) / historical_sales_count) * 100, 2)
        else:
            ratio_5 = 0.0
        results['indicator_5'] = {
            'name': '二次限流动销品占比',
            'value': ratio_5,
            'unit': '%'
        }

        # 指标6：日均单量（近7日）
        volume_6 = get_recent_sales_volume(sales_table_name, days=7, end_date=end_date)
        results['indicator_6'] = {
            'name': '日均单量（近7日）',
            'value': volume_6,
            'unit': '个'
        }

        # 指标7：过程数据展示
        process_data = {
            'low_risk_active_count': len(active_goods_ids),  # 无风险active数量
            'high_risk_active_count': len(at_risk_goods_ids),  # 有风险active数量
            'restricted_data_count': len(restricted_goods_ids),  # 限流数据数量（限流目录中所有xlsx文件的goods_id数量）
            'unpriced_data_count': len(unpriced_goods_ids),  # 未核价数据数量（未核价目录中所有xlsx文件的goods_id数量）
            'restricted_sales_count': len(restricted_sales_goods)  # 历史动销被限流数量
        }
        results['indicator_7'] = {
            'name': '过程数据展示',
            'value': process_data,
            'unit': ''
        }

        # 指标8和9：GMV图表
        # 只有在非缓存模式（use_cache=False）时才生成图表
        if not use_cache:
            print("[图表生成] 非缓存模式，生成GMV图表")
            df_30day = get_gmv_data(sales_table_name, days=30, end_date=end_date)
            
            # 指标8：近30天GMV图
            chart_8 = generate_gmv_chart(df_30day, '近30天GMV图', 'skyblue', 'red')
            results['indicator_8'] = {
                'name': '近30天GMV图',
                'value': chart_8,
                'unit': ''
            }

            # 指标9：日均GMV图（使用30天数据的后7天）
            df_7day = df_30day.tail(7).reset_index(drop=True) if len(df_30day) >= 7 else df_30day
            chart_9 = generate_gmv_chart(df_7day, '日均GMV图（近7日）', 'lightgreen', 'orange')
            results['indicator_9'] = {
                'name': '日均GMV图',
                'value': chart_9,
                'unit': ''
            }
        else:
            # 缓存模式下不生成图表
            print("[图表跳过] 缓存模式，跳过GMV图表生成")
            results['indicator_8'] = {
                'name': '近30天GMV图',
                'value': None,
                'unit': ''
            }
            results['indicator_9'] = {
                'name': '日均GMV图',
                'value': None,
                'unit': ''
            }

        # 计算运行时间
        end_time = time.time()
        analysis_time = round(end_time - start_time, 2)
        
        # 保存缓存（无论use_cache是True还是False都保存，以便下次使用）
        save_indicator_cache(current_table, target_date_str, results, analysis_time)

        return jsonify({
            'success': True,
            'data': results,
            'analysis_time': analysis_time,
            'from_cache': False
        })

    except Exception as e:
        # 即使出错也计算运行时间
        end_time = time.time()
        analysis_time = round(end_time - start_time, 2)
        
        return jsonify({
            'success': False,
            'error': f'计算指标时出错: {str(e)}',
            'analysis_time': analysis_time
        }), 500


def configure_directories():
    """
    配置Excel文件目录
    """
    try:
        data = request.json
        unpriced_dir = data.get('unpriced_data_dir', '').strip()
        restricted_dir = data.get('traffic_restricted_data_dir', '').strip()

        if not unpriced_dir or not restricted_dir:
            return jsonify({
                'success': False,
                'error': '未核价数据目录和限流数据目录都不能为空'
            }), 400

        # 验证目录
        is_valid, error_msg, file_counts = validate_excel_directories(unpriced_dir, restricted_dir)

        if not is_valid:
            return jsonify({
                'success': False,
                'error': error_msg
            }), 400

        # 保存配置
        config_data = {
            'unpriced_data_dir': unpriced_dir,
            'traffic_restricted_data_dir': restricted_dir
        }

        if save_indicator_config(config_data):
            return jsonify({
                'success': True,
                'message': '配置成功',
                'file_counts': file_counts
            })
        else:
            return jsonify({
                'success': False,
                'error': '保存配置失败'
            }), 500

    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'配置过程中出错: {str(e)}'
        }), 500


def get_indicator_config():
    """
    获取当前配置
    """
    try:
        config = load_indicator_config()
        return jsonify({
            'success': True,
            'data': config
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'获取配置时出错: {str(e)}'
        }), 500


def save_indicator_data_to_excel(target_date=None):
    """
    保存指标数据到Excel文件
    文件名为"指标体系数据.xlsx"，每个sheet为表名
    始终追加新行，不覆盖已有数据
    参数:
        target_date: 目标日期（字符串格式 'YYYY-MM-DD'），如果为None则使用昨天（东欧时间-1天）
    """
    try:
        # 处理日期参数
        if target_date is None:
            record_date = (get_eastern_europe_time() - timedelta(days=1)).date()
        else:
            if isinstance(target_date, str):
                record_date = datetime.strptime(target_date, '%Y-%m-%d').date()
            elif isinstance(target_date, datetime):
                record_date = target_date.date()
            else:
                record_date = target_date
        
        # 获取当前指标数据
        current_table = get_current_table()
        sales_table_name = f"{current_table}_Sales"
        
        # 获取配置
        config = load_indicator_config()
        unpriced_dir = config.get('unpriced_data_dir', '')
        restricted_dir = config.get('traffic_restricted_data_dir', '')
        
        unpriced_goods_ids = set()
        restricted_goods_ids = set()
        
        if unpriced_dir and restricted_dir:
            unpriced_goods_ids, restricted_goods_ids = get_excel_data(unpriced_dir, restricted_dir)
        
        # 获取基础数据
        active_goods_ids, at_risk_goods_ids = get_active_products_data(current_table)
        all_active_goods = active_goods_ids.union(at_risk_goods_ids)
        sales_goods_ids = get_sales_data(current_table, sales_table_name, end_date=record_date)
        
        excluded_goods = unpriced_goods_ids.union(restricted_goods_ids)
        non_restricted_active_goods = all_active_goods - excluded_goods
        restricted_sales_goods = sales_goods_ids.intersection(restricted_goods_ids)
        
        # 计算指标数据
        count_1 = len(non_restricted_active_goods)  # 本周非限流在售
        restricted_count = len(restricted_goods_ids)
        denominator = count_1 + restricted_count
        ratio_2 = round((restricted_count / denominator) * 100, 2) if denominator > 0 else 0.0  # 二次限流占比
        count_3 = len(sales_goods_ids)  # 本周动销品数
        active_sales_goods = non_restricted_active_goods.intersection(sales_goods_ids)
        count_4 = len(active_sales_goods)  # 本周在售动销品数
        historical_sales_count = len(sales_goods_ids)
        ratio_5 = round((len(restricted_sales_goods) / historical_sales_count) * 100, 2) if historical_sales_count > 0 else 0.0  # 二次限流动销品占比
        volume_6 = get_recent_sales_volume(sales_table_name, days=7, end_date=record_date)  # 日均单量（近7日）
        
        # 过程数据
        process_data = {
            'low_risk_active_count': len(active_goods_ids),
            'high_risk_active_count': len(at_risk_goods_ids),
            'restricted_data_count': len(restricted_goods_ids),
            'unpriced_data_count': len(unpriced_goods_ids),
            'restricted_sales_count': len(restricted_sales_goods)
        }
        
        # 记录日期已在函数开始处处理（使用target_date参数）
        record_date_str = record_date.strftime('%Y-%m-%d')
        
        # Excel文件路径（当前目录）
        excel_file = '指标体系数据.xlsx'
        
        # 定义列名
        columns = [
            '记录日期',
            '上周非限流在售',
            '本周非限流在售',
            '增长率（非限流在售）',
            '二次限流占比',
            '上周动销品数',
            '本周动销品数',
            '上周畅销品数',
            '本周畅销品数',
            '上周在售动销品数',
            '本周在售动销品数',
            '增长率（在售动销品）',
            '二次限流动销品占比',
            '日均单量（近7日）',
            '无风险active数量',
            '有风险active数量',
            '限流数据数量',
            '未核价数据数量',
            '历史动销被限流数量'
        ]
        
        # 读取或创建Excel文件
        if os.path.exists(excel_file):
            # 文件存在，读取所有sheet
            excel_data = pd.read_excel(excel_file, sheet_name=None, engine='openpyxl')
        else:
            # 文件不存在，创建空字典
            excel_data = {}
        
        # 获取或创建当前表的sheet
        if current_table in excel_data:
            df = excel_data[current_table]
        else:
            # 创建新sheet
            df = pd.DataFrame(columns=columns)
            excel_data[current_table] = df
        
        # 确保列存在
        for col in columns:
            if col not in df.columns:
                df[col] = None
        
        # 始终追加新行，不覆盖已有数据
        row_idx = len(df)
        # 获取最后一行数据（用于计算"上周"数据）
        if len(df) > 0:
            last_week_non_restricted = df.iloc[-1]['本周非限流在售'] if pd.notna(df.iloc[-1]['本周非限流在售']) else 0
            last_week_sales = df.iloc[-1]['本周动销品数'] if pd.notna(df.iloc[-1]['本周动销品数']) else 0
            last_week_active_sales = df.iloc[-1]['本周在售动销品数'] if pd.notna(df.iloc[-1]['本周在售动销品数']) else 0
        else:
            last_week_non_restricted = 0
            last_week_sales = 0
            last_week_active_sales = 0
        
        # 计算增长率
        if last_week_non_restricted > 0:
            growth_rate_1 = round(((count_1 - last_week_non_restricted) / last_week_non_restricted) * 100, 2)
        else:
            growth_rate_1 = 0.0
        
        if last_week_active_sales > 0:
            growth_rate_2 = round(((count_4 - last_week_active_sales) / last_week_active_sales) * 100, 2)
        else:
            growth_rate_2 = 0.0
        
        # 准备新行数据（使用列表按顺序存储，因为有两个"增长率"列）
        # 注意：百分比值在写入DataFrame时转换为小数（除以100），以便Excel百分比格式正确显示
        new_row_data = [
            record_date_str,  # 记录日期
            last_week_non_restricted,  # 上周非限流在售
            count_1,  # 本周非限流在售
            growth_rate_1 / 100.0,  # 增长率 - 转换为小数
            ratio_2 / 100.0,  # 二次限流占比 - 转换为小数
            last_week_sales,  # 上周动销品数
            count_3,  # 本周动销品数
            None,  # 上周畅销品数
            None,  # 本周畅销品数
            last_week_active_sales,  # 上周在售动销品数
            count_4,  # 本周在售动销品数
            growth_rate_2 / 100.0,  # 增长率（第二个）- 转换为小数
            ratio_5 / 100.0,  # 二次限流动销品占比 - 转换为小数
            volume_6,  # 日均单量（近7日）
            process_data['low_risk_active_count'],  # 无风险active数量
            process_data['high_risk_active_count'],  # 有风险active数量
            process_data['restricted_data_count'],  # 限流数据数量
            process_data['unpriced_data_count'],  # 未核价数据数量
            process_data['restricted_sales_count']  # 历史动销被限流数量
        ]
        
        # 如果行索引在现有数据范围内，更新行；否则追加新行
        if row_idx < len(df):
            # 更新现有行
            for i, col in enumerate(columns):
                df.at[row_idx, col] = new_row_data[i]
        else:
            # 追加新行（使用列表创建DataFrame，按列顺序）
            new_df = pd.DataFrame([new_row_data], columns=columns)
            df = pd.concat([df, new_df], ignore_index=True)
        
        # 更新excel_data
        excel_data[current_table] = df
        
        # 保存到Excel文件（使用openpyxl引擎）
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            for sheet_name, sheet_df in excel_data.items():
                sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 使用openpyxl设置百分比格式
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import NamedStyle
            import openpyxl.utils
            
            wb = load_workbook(excel_file)
            if current_table in wb.sheetnames:
                ws = wb[current_table]
                
                # 定义百分比列的索引（从0开始）
                # 列索引：增长率（非限流在售）(3), 二次限流占比(4), 增长率（在售动销品）(11), 二次限流动销品占比(12)
                percentage_cols = [3, 4, 11, 12]  # D列(增长率（非限流在售）), E列(二次限流占比), L列(增长率（在售动销品）), M列(二次限流动销品占比)
                
                # 设置百分比格式样式
                percentage_style = NamedStyle(name="percentage_style", number_format='0.00%')
                
                # 为所有数据行设置百分比格式（值已经在DataFrame中转换为小数）
                # 新行的值已经是小数格式，旧行的值也是小数格式（之前已经转换过）
                for col_idx in percentage_cols:
                    col_letter = openpyxl.utils.get_column_letter(col_idx + 1)  # +1因为Excel列从1开始
                    # 从第2行开始（第1行是表头）
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws[f'{col_letter}{row_idx}']
                        if cell.value is not None:
                            cell.number_format = '0.00%'
                
                wb.save(excel_file)
                wb.close()
        except Exception as format_error:
            print(f"设置百分比格式时出错（数据仍会保存）: {format_error}")
            # 即使格式化失败，数据仍然保存成功
        
        return jsonify({
            'success': True,
            'message': '保存成功'
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'保存指标数据时出错: {str(e)}'
        }), 500


# ===== 批量操作支持函数 =====

def get_excel_data_for_table(unpriced_dir, restricted_dir, table_name):
    """
    获取指定表的Excel文件中的goods_id数据
    参数:
        unpriced_dir: 未核价数据目录
        restricted_dir: 限流数据目录
        table_name: 表名（站点名称）
    返回: (unpriced_goods_ids, restricted_goods_ids)
    """
    try:
        unpriced_goods_ids, restricted_goods_ids = read_excel_files_to_goods_ids(unpriced_dir, restricted_dir, table_name)
        return unpriced_goods_ids, restricted_goods_ids
    except Exception as e:
        print(f"获取Excel数据出错: {e}")
        import traceback
        traceback.print_exc()
        return set(), set()


def get_active_products_data_for_table(table_name):
    """
    获取指定表的商品表中在售商品数据
    参数:
        table_name: 表名
    返回: (active_goods_ids, at_risk_goods_ids)
    """
    _, _, _, product_config = get_db_config()

    try:
        conn = get_db_connection(product_config)
        cursor = conn.cursor()

        # 获取Active状态的商品
        active_query = f"""
        SELECT DISTINCT goods_id
        FROM `{table_name}`
        WHERE detail_status = 'Active'
        """
        cursor.execute(active_query)
        active_results = cursor.fetchall()
        active_goods_ids = set()
        for row in active_results:
            normalized_id = normalize_goods_id(row[0])
            if normalized_id:
                active_goods_ids.add(normalized_id)

        # 获取At Risk状态的商品
        at_risk_query = f"""
        SELECT DISTINCT goods_id
        FROM `{table_name}`
        WHERE detail_status = 'At Risk'
        """
        cursor.execute(at_risk_query)
        at_risk_results = cursor.fetchall()
        at_risk_goods_ids = set()
        for row in at_risk_results:
            normalized_id = normalize_goods_id(row[0])
            if normalized_id:
                at_risk_goods_ids.add(normalized_id)

        cursor.close()
        conn.close()

        return active_goods_ids, at_risk_goods_ids

    except Exception as e:
        print(f"获取商品数据出错: {e}")
        return set(), set()


def get_sales_data_for_table(table_name, sales_table_name, end_date=None):
    """
    获取指定表的销售表中动销商品数据
    参数:
        table_name: 表名
        sales_table_name: 销售表名
        end_date: 结束日期
    返回: sales_goods_ids (set)
    """
    _, sales_config, _, _ = get_db_config()

    try:
        conn = get_db_connection(sales_config)
        cursor = conn.cursor()

        if end_date is None:
            sales_query = f"""
            SELECT DISTINCT goods_id
            FROM `{sales_table_name}`
            """
            cursor.execute(sales_query)
        else:
            if isinstance(end_date, datetime):
                end_date = end_date.date()
            elif isinstance(end_date, str):
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            
            end_date_str = end_date.strftime('%Y-%m-%d')
            sales_query = f"""
            SELECT DISTINCT goods_id
            FROM `{sales_table_name}`
            WHERE date_label <= %s
            """
            cursor.execute(sales_query, (end_date_str,))
        
        sales_results = cursor.fetchall()
        sales_goods_ids = set()
        for row in sales_results:
            normalized_id = normalize_goods_id(row[0])
            if normalized_id:
                sales_goods_ids.add(normalized_id)

        cursor.close()
        conn.close()

        return sales_goods_ids

    except Exception as e:
        print(f"获取销售数据出错: {e}")
        return set()


def get_recent_sales_volume_for_table(sales_table_name, days=7, end_date=None):
    """
    获取指定表的近N天日均单量
    """
    _, sales_config, _, _ = get_db_config()

    try:
        conn = get_db_connection(sales_config)
        cursor = conn.cursor()

        if end_date is None:
            end_date = (get_eastern_europe_time() - timedelta(days=1)).date()
        elif isinstance(end_date, datetime):
            end_date = end_date.date()
        elif isinstance(end_date, str):
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        start_date = end_date - timedelta(days=days-1)

        start_date_str = start_date.strftime('%Y-%m-%d')
        end_date_str = end_date.strftime('%Y-%m-%d')

        volume_query = f"""
        SELECT SUM(`Units ordered`) as total_volume
        FROM `{sales_table_name}`
        WHERE date_label >= %s AND date_label <= %s
        """
        cursor.execute(volume_query, (start_date_str, end_date_str))
        result = cursor.fetchone()
        total_volume = result[0] if result[0] else 0

        avg_daily_volume = round(total_volume / days, 2)

        cursor.close()
        conn.close()

        return avg_daily_volume

    except Exception as e:
        print(f"获取销售量数据出错: {e}")
        return 0.0


def indicator_calculation_for_table(table_name, target_date=None, use_cache=True):
    """
    针对指定表名计算指标（用于批量操作）
    参数:
        table_name: 要计算的表名
        target_date: 目标日期
        use_cache: 是否使用缓存
    返回: 计算结果字典（不是jsonify对象）
    """
    import time
    start_time = time.time()
    
    try:
        # 处理日期参数
        if target_date is None:
            end_date = (get_eastern_europe_time() - timedelta(days=1)).date()
        else:
            if isinstance(target_date, str):
                end_date = datetime.strptime(target_date, '%Y-%m-%d').date()
            elif isinstance(target_date, datetime):
                end_date = target_date.date()
            else:
                end_date = target_date
        
        target_date_str = end_date.strftime('%Y-%m-%d')
        sales_table_name = f"{table_name}_Sales"
        
        # 如果使用缓存，先检查缓存
        if use_cache:
            cache_data = load_indicator_cache(table_name, target_date_str)
            if cache_data:
                end_time = time.time()
                return {
                    'success': True,
                    'data': cache_data.get('results', {}),
                    'analysis_time': round(end_time - start_time, 2),
                    'from_cache': True
                }
        
        results = {}

        # 获取配置
        config = load_indicator_config()
        unpriced_dir = config.get('unpriced_data_dir', '')
        restricted_dir = config.get('traffic_restricted_data_dir', '')
        
        unpriced_goods_ids = set()
        restricted_goods_ids = set()
        
        if unpriced_dir and restricted_dir:
            unpriced_goods_ids, restricted_goods_ids = get_excel_data_for_table(unpriced_dir, restricted_dir, table_name)

        # 获取商品表数据
        active_goods_ids, at_risk_goods_ids = get_active_products_data_for_table(table_name)
        all_active_goods = active_goods_ids.union(at_risk_goods_ids)
        
        # 获取销售表数据
        sales_goods_ids = get_sales_data_for_table(table_name, sales_table_name, end_date=end_date)
        
        # 计算公共数据
        excluded_goods = unpriced_goods_ids.union(restricted_goods_ids)
        non_restricted_active_goods = all_active_goods - excluded_goods
        restricted_sales_goods = sales_goods_ids.intersection(restricted_goods_ids)

        # 指标1：非限流在售商品数量
        count_1 = len(non_restricted_active_goods)
        results['indicator_1'] = {
            'name': '非限流在售商品数量',
            'value': count_1,
            'unit': '个'
        }

        # 指标2：二次限流占比
        restricted_count = len(restricted_goods_ids)
        denominator = count_1 + restricted_count
        if denominator > 0:
            ratio_2 = round((restricted_count / denominator) * 100, 2)
        else:
            ratio_2 = 0.0
        results['indicator_2'] = {
            'name': '二次限流占比',
            'value': ratio_2,
            'unit': '%'
        }

        # 指标3：历史动销品数量
        count_3 = len(sales_goods_ids)
        results['indicator_3'] = {
            'name': '历史动销品数量',
            'value': count_3,
            'unit': '个'
        }

        # 指标4：在售动销品数量
        active_sales_goods = non_restricted_active_goods.intersection(sales_goods_ids)
        count_4 = len(active_sales_goods)
        results['indicator_4'] = {
            'name': '在售动销品数量',
            'value': count_4,
            'unit': '个',
            'goods_ids': sorted(list(active_sales_goods))
        }

        # 指标5：二次限流动销品占比
        historical_sales_count = len(sales_goods_ids)
        if historical_sales_count > 0:
            ratio_5 = round((len(restricted_sales_goods) / historical_sales_count) * 100, 2)
        else:
            ratio_5 = 0.0
        results['indicator_5'] = {
            'name': '二次限流动销品占比',
            'value': ratio_5,
            'unit': '%'
        }

        # 指标6：日均单量（近7日）
        volume_6 = get_recent_sales_volume_for_table(sales_table_name, days=7, end_date=end_date)
        results['indicator_6'] = {
            'name': '日均单量（近7日）',
            'value': volume_6,
            'unit': '个'
        }

        # 指标7：过程数据展示
        process_data = {
            'low_risk_active_count': len(active_goods_ids),
            'high_risk_active_count': len(at_risk_goods_ids),
            'restricted_data_count': len(restricted_goods_ids),
            'unpriced_data_count': len(unpriced_goods_ids),
            'restricted_sales_count': len(restricted_sales_goods)
        }
        results['indicator_7'] = {
            'name': '过程数据展示',
            'value': process_data,
            'unit': ''
        }

        # 计算运行时间
        end_time = time.time()
        analysis_time = round(end_time - start_time, 2)
        
        # 保存缓存
        save_indicator_cache(table_name, target_date_str, results, analysis_time)

        return {
            'success': True,
            'data': results,
            'analysis_time': analysis_time,
            'from_cache': False
        }

    except Exception as e:
        end_time = time.time()
        analysis_time = round(end_time - start_time, 2)
        
        return {
            'success': False,
            'error': f'计算指标时出错: {str(e)}',
            'analysis_time': analysis_time
        }


def save_indicator_data_to_excel_for_table(table_name, target_date=None):
    """
    针对指定表名保存指标数据到Excel（用于批量操作）
    参数:
        table_name: 要保存的表名
        target_date: 目标日期
    返回: 结果字典（不是jsonify对象）
    """
    try:
        # 处理日期参数
        if target_date is None:
            record_date = (get_eastern_europe_time() - timedelta(days=1)).date()
        else:
            if isinstance(target_date, str):
                record_date = datetime.strptime(target_date, '%Y-%m-%d').date()
            elif isinstance(target_date, datetime):
                record_date = target_date.date()
            else:
                record_date = target_date
        
        sales_table_name = f"{table_name}_Sales"
        
        # 获取配置
        config = load_indicator_config()
        unpriced_dir = config.get('unpriced_data_dir', '')
        restricted_dir = config.get('traffic_restricted_data_dir', '')
        
        unpriced_goods_ids = set()
        restricted_goods_ids = set()
        
        if unpriced_dir and restricted_dir:
            unpriced_goods_ids, restricted_goods_ids = get_excel_data_for_table(unpriced_dir, restricted_dir, table_name)
        
        # 获取基础数据
        active_goods_ids, at_risk_goods_ids = get_active_products_data_for_table(table_name)
        all_active_goods = active_goods_ids.union(at_risk_goods_ids)
        sales_goods_ids = get_sales_data_for_table(table_name, sales_table_name, end_date=record_date)
        
        excluded_goods = unpriced_goods_ids.union(restricted_goods_ids)
        non_restricted_active_goods = all_active_goods - excluded_goods
        restricted_sales_goods = sales_goods_ids.intersection(restricted_goods_ids)
        
        # 计算指标数据
        count_1 = len(non_restricted_active_goods)
        restricted_count = len(restricted_goods_ids)
        denominator = count_1 + restricted_count
        ratio_2 = round((restricted_count / denominator) * 100, 2) if denominator > 0 else 0.0
        count_3 = len(sales_goods_ids)
        active_sales_goods = non_restricted_active_goods.intersection(sales_goods_ids)
        count_4 = len(active_sales_goods)
        historical_sales_count = len(sales_goods_ids)
        ratio_5 = round((len(restricted_sales_goods) / historical_sales_count) * 100, 2) if historical_sales_count > 0 else 0.0
        volume_6 = get_recent_sales_volume_for_table(sales_table_name, days=7, end_date=record_date)
        
        process_data = {
            'low_risk_active_count': len(active_goods_ids),
            'high_risk_active_count': len(at_risk_goods_ids),
            'restricted_data_count': len(restricted_goods_ids),
            'unpriced_data_count': len(unpriced_goods_ids),
            'restricted_sales_count': len(restricted_sales_goods)
        }
        
        record_date_str = record_date.strftime('%Y-%m-%d')
        excel_file = '指标体系数据.xlsx'
        
        columns = [
            '记录日期', '上周非限流在售', '本周非限流在售', '增长率（非限流在售）',
            '二次限流占比', '上周动销品数', '本周动销品数', '上周畅销品数',
            '本周畅销品数', '上周在售动销品数', '本周在售动销品数', '增长率（在售动销品）',
            '二次限流动销品占比', '日均单量（近7日）', '无风险active数量',
            '有风险active数量', '限流数据数量', '未核价数据数量', '历史动销被限流数量'
        ]
        
        # 读取或创建Excel文件
        if os.path.exists(excel_file):
            excel_data = pd.read_excel(excel_file, sheet_name=None, engine='openpyxl')
        else:
            excel_data = {}
        
        if table_name in excel_data:
            df = excel_data[table_name]
        else:
            df = pd.DataFrame(columns=columns)
            excel_data[table_name] = df
        
        for col in columns:
            if col not in df.columns:
                df[col] = None
        
        row_idx = len(df)
        if len(df) > 0:
            last_week_non_restricted = df.iloc[-1]['本周非限流在售'] if pd.notna(df.iloc[-1]['本周非限流在售']) else 0
            last_week_sales = df.iloc[-1]['本周动销品数'] if pd.notna(df.iloc[-1]['本周动销品数']) else 0
            last_week_active_sales = df.iloc[-1]['本周在售动销品数'] if pd.notna(df.iloc[-1]['本周在售动销品数']) else 0
        else:
            last_week_non_restricted = 0
            last_week_sales = 0
            last_week_active_sales = 0
        
        if last_week_non_restricted > 0:
            growth_rate_1 = round(((count_1 - last_week_non_restricted) / last_week_non_restricted) * 100, 2)
        else:
            growth_rate_1 = 0.0
        
        if last_week_active_sales > 0:
            growth_rate_2 = round(((count_4 - last_week_active_sales) / last_week_active_sales) * 100, 2)
        else:
            growth_rate_2 = 0.0
        
        new_row_data = [
            record_date_str, last_week_non_restricted, count_1, growth_rate_1 / 100.0,
            ratio_2 / 100.0, last_week_sales, count_3, None, None,
            last_week_active_sales, count_4, growth_rate_2 / 100.0, ratio_5 / 100.0,
            volume_6, process_data['low_risk_active_count'], process_data['high_risk_active_count'],
            process_data['restricted_data_count'], process_data['unpriced_data_count'],
            process_data['restricted_sales_count']
        ]
        
        if row_idx < len(df):
            for i, col in enumerate(columns):
                df.at[row_idx, col] = new_row_data[i]
        else:
            new_df = pd.DataFrame([new_row_data], columns=columns)
            df = pd.concat([df, new_df], ignore_index=True)
        
        excel_data[table_name] = df
        
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            for sheet_name, sheet_df in excel_data.items():
                sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 设置百分比格式
        try:
            from openpyxl import load_workbook
            import openpyxl.utils
            
            wb = load_workbook(excel_file)
            if table_name in wb.sheetnames:
                ws = wb[table_name]
                percentage_cols = [3, 4, 11, 12]
                
                for col_idx in percentage_cols:
                    col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws[f'{col_letter}{row_idx}']
                        if cell.value is not None:
                            cell.number_format = '0.00%'
                
                wb.save(excel_file)
                wb.close()
        except Exception as format_error:
            print(f"设置百分比格式时出错: {format_error}")
        
        return {
            'success': True,
            'message': '保存成功'
        }
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {
            'success': False,
            'error': f'保存指标数据时出错: {str(e)}'
        }